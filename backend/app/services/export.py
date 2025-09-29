"""
Export Service for Financial Document Analyzer
Handles export functionality for reports, analysis results, and dashboard data
"""

import os
import io
import json
import csv
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import logging
from pathlib import Path
import zipfile
import tempfile

# PDF and Excel export dependencies
try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False
    logging.warning("ReportLab not available. PDF export will be limited.")

try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False
    logging.warning("Pandas/OpenPyXL not available. Excel export will be limited.")

from app.core.database import get_mongodb_client
from app.models.schemas import LogLevel, LogCategory, LogAction

# Set up logging
logger = logging.getLogger(__name__)

class ExportService:
    """Service for exporting analysis results and reports"""
    
    def __init__(self):
        self.mongodb_client = get_mongodb_client()
        self.export_dir = Path("exports")
        self.export_dir.mkdir(exist_ok=True)
        
        # Initialize logging service
        from app.core.logging import get_logging_service
        self.logging_service = get_logging_service()
    
    def export_analysis_to_pdf(
        self,
        analysis_data: Dict[str, Any],
        user_id: str,
        filename: Optional[str] = None
    ) -> str:
        """Export analysis results to PDF format"""
        try:
            if not REPORTLAB_AVAILABLE:
                raise Exception("ReportLab not available for PDF export")
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analysis_report_{timestamp}.pdf"
            
            filepath = self.export_dir / filename
            
            # Create PDF document
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            story.append(Paragraph("Financial Document Analysis Report", title_style))
            story.append(Spacer(1, 12))
            
            # Document information
            doc_info = analysis_data.get('document_info', {})
            if doc_info:
                story.append(Paragraph("Document Information", styles['Heading2']))
                info_data = [
                    ['Document Name:', doc_info.get('filename', 'N/A')],
                    ['Upload Date:', doc_info.get('upload_date', 'N/A')],
                    ['File Size:', f"{doc_info.get('file_size', 0)} bytes"],
                    ['Analysis Date:', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
                ]
                info_table = Table(info_data, colWidths=[2*inch, 4*inch])
                info_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                    ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                    ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 0), (-1, -1), 10),
                    ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
                    ('BACKGROUND', (1, 0), (1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                story.append(info_table)
                story.append(Spacer(1, 20))
            
            # Executive Summary
            summary = analysis_data.get('executive_summary', '')
            if summary:
                story.append(Paragraph("Executive Summary", styles['Heading2']))
                story.append(Paragraph(summary, styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Financial Analysis
            financial_analysis = analysis_data.get('financial_analysis', {})
            if financial_analysis:
                story.append(Paragraph("Financial Analysis", styles['Heading2']))
                
                # Key metrics table
                metrics = financial_analysis.get('key_metrics', {})
                if metrics:
                    story.append(Paragraph("Key Financial Metrics", styles['Heading3']))
                    metrics_data = [['Metric', 'Value', 'Analysis']]
                    for metric, value in metrics.items():
                        metrics_data.append([metric, str(value), ''])
                    
                    metrics_table = Table(metrics_data, colWidths=[2*inch, 1.5*inch, 2.5*inch])
                    metrics_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 12),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    story.append(metrics_table)
                    story.append(Spacer(1, 20))
            
            # Risk Assessment
            risk_assessment = analysis_data.get('risk_assessment', {})
            if risk_assessment:
                story.append(Paragraph("Risk Assessment", styles['Heading2']))
                risk_level = risk_assessment.get('risk_level', 'N/A')
                risk_factors = risk_assessment.get('risk_factors', [])
                
                story.append(Paragraph(f"Overall Risk Level: {risk_level}", styles['Normal']))
                if risk_factors:
                    story.append(Paragraph("Risk Factors:", styles['Heading3']))
                    for factor in risk_factors:
                        story.append(Paragraph(f"• {factor}", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Investment Recommendations
            recommendations = analysis_data.get('investment_recommendations', {})
            if recommendations:
                story.append(Paragraph("Investment Recommendations", styles['Heading2']))
                recommendation = recommendations.get('recommendation', '')
                confidence = recommendations.get('confidence', 'N/A')
                
                story.append(Paragraph(f"Recommendation: {recommendation}", styles['Normal']))
                story.append(Paragraph(f"Confidence Level: {confidence}", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Detailed Analysis
            detailed_analysis = analysis_data.get('detailed_analysis', '')
            if detailed_analysis:
                story.append(Paragraph("Detailed Analysis", styles['Heading2']))
                story.append(Paragraph(detailed_analysis, styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Footer
            story.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                alignment=TA_CENTER,
                textColor=colors.grey
            )
            story.append(Paragraph(f"Generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", footer_style))
            story.append(Paragraph("Financial Document Analyzer", footer_style))
            
            # Build PDF
            doc.build(story)
            
            # Log export activity
            self.logging_service.log_activity(
                level=LogLevel.INFO,
                category=LogCategory.EXPORT,
                action=LogAction.EXPORT,
                message=f"Analysis exported to PDF: {filename}",
                user_id=user_id,
                details={
                    "export_type": "pdf",
                    "filename": filename,
                    "file_size": filepath.stat().st_size
                }
            )
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting analysis to PDF: {e}")
            self.logging_service.log_activity(
                level=LogLevel.ERROR,
                category=LogCategory.EXPORT,
                action=LogAction.EXPORT,
                message=f"Failed to export analysis to PDF: {str(e)}",
                user_id=user_id
            )
            raise
    
    def export_analysis_to_excel(
        self,
        analysis_data: Dict[str, Any],
        user_id: str,
        filename: Optional[str] = None
    ) -> str:
        """Export analysis results to Excel format"""
        try:
            if not PANDAS_AVAILABLE:
                raise Exception("Pandas/OpenPyXL not available for Excel export")
            
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"analysis_report_{timestamp}.xlsx"
            
            filepath = self.export_dir / filename
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Document Information Sheet
            doc_sheet = wb.create_sheet("Document Information")
            doc_info = analysis_data.get('document_info', {})
            doc_data = [
                ['Document Name', doc_info.get('filename', 'N/A')],
                ['Upload Date', doc_info.get('upload_date', 'N/A')],
                ['File Size (bytes)', doc_info.get('file_size', 0)],
                ['Analysis Date', datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
            ]
            
            for row in doc_data:
                doc_sheet.append(row)
            
            # Style the document info sheet
            self._style_excel_sheet(doc_sheet, "Document Information")
            
            # Executive Summary Sheet
            summary_sheet = wb.create_sheet("Executive Summary")
            summary = analysis_data.get('executive_summary', '')
            summary_sheet.append(['Executive Summary'])
            summary_sheet.append([summary])
            self._style_excel_sheet(summary_sheet, "Executive Summary")
            
            # Financial Analysis Sheet
            financial_analysis = analysis_data.get('financial_analysis', {})
            if financial_analysis:
                fin_sheet = wb.create_sheet("Financial Analysis")
                
                # Key metrics
                metrics = financial_analysis.get('key_metrics', {})
                if metrics:
                    fin_sheet.append(['Key Financial Metrics'])
                    fin_sheet.append(['Metric', 'Value', 'Analysis'])
                    
                    for metric, value in metrics.items():
                        fin_sheet.append([metric, value, ''])
                    
                    self._style_excel_sheet(fin_sheet, "Financial Analysis")
            
            # Risk Assessment Sheet
            risk_assessment = analysis_data.get('risk_assessment', {})
            if risk_assessment:
                risk_sheet = wb.create_sheet("Risk Assessment")
                
                risk_sheet.append(['Risk Assessment'])
                risk_sheet.append(['Overall Risk Level', risk_assessment.get('risk_level', 'N/A')])
                risk_sheet.append([''])
                risk_sheet.append(['Risk Factors'])
                
                risk_factors = risk_assessment.get('risk_factors', [])
                for factor in risk_factors:
                    risk_sheet.append([factor])
                
                self._style_excel_sheet(risk_sheet, "Risk Assessment")
            
            # Investment Recommendations Sheet
            recommendations = analysis_data.get('investment_recommendations', {})
            if recommendations:
                rec_sheet = wb.create_sheet("Investment Recommendations")
                
                rec_sheet.append(['Investment Recommendations'])
                rec_sheet.append(['Recommendation', recommendations.get('recommendation', '')])
                rec_sheet.append(['Confidence Level', recommendations.get('confidence', 'N/A')])
                
                self._style_excel_sheet(rec_sheet, "Investment Recommendations")
            
            # Detailed Analysis Sheet
            detailed_analysis = analysis_data.get('detailed_analysis', '')
            if detailed_analysis:
                detail_sheet = wb.create_sheet("Detailed Analysis")
                detail_sheet.append(['Detailed Analysis'])
                detail_sheet.append([detailed_analysis])
                self._style_excel_sheet(detail_sheet, "Detailed Analysis")
            
            # Save workbook
            wb.save(str(filepath))
            
            # Log export activity
            self.logging_service.log_activity(
                level=LogLevel.INFO,
                category=LogCategory.EXPORT,
                action=LogAction.EXPORT,
                message=f"Analysis exported to Excel: {filename}",
                user_id=user_id,
                details={
                    "export_type": "excel",
                    "filename": filename,
                    "file_size": filepath.stat().st_size
                }
            )
            
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Error exporting analysis to Excel: {e}")
            self.logging_service.log_activity(
                level=LogLevel.ERROR,
                category=LogCategory.EXPORT,
                action=LogAction.EXPORT,
                message=f"Failed to export analysis to Excel: {str(e)}",
                user_id=user_id
            )
            raise
    
    def export_dashboard_data(
        self,
        dashboard_data: Dict[str, Any],
        user_id: str,
        format_type: str = "excel",
        filename: Optional[str] = None
    ) -> str:
        """Export dashboard data to specified format"""
        try:
            if not filename:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"dashboard_data_{timestamp}.{format_type}"
            
            filepath = self.export_dir / filename
            
            if format_type.lower() == "excel":
                return self._export_dashboard_to_excel(dashboard_data, filepath, user_id)
            elif format_type.lower() == "csv":
                return self._export_dashboard_to_csv(dashboard_data, filepath, user_id)
            elif format_type.lower() == "json":
                return self._export_dashboard_to_json(dashboard_data, filepath, user_id)
            else:
                raise ValueError(f"Unsupported export format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error exporting dashboard data: {e}")
            raise
    
    def export_user_history(
        self,
        user_id: str,
        format_type: str = "excel",
        days: int = 30
    ) -> str:
        """Export user's analysis history"""
        try:
            if not self.mongodb_client.db:
                raise Exception("MongoDB not available")
            
            # Get user's analysis history
            end_date = datetime.utcnow()
            start_date = datetime(end_date.year, end_date.month, end_date.day - days)
            
            history = list(self.mongodb_client.db['document_analyses'].find({
                'user_id': user_id,
                'created_at': {'$gte': start_date, '$lte': end_date}
            }).sort('created_at', -1))
            
            if not history:
                raise Exception("No analysis history found for the specified period")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"user_history_{user_id}_{timestamp}.{format_type}"
            filepath = self.export_dir / filename
            
            if format_type.lower() == "excel":
                return self._export_history_to_excel(history, filepath, user_id)
            elif format_type.lower() == "csv":
                return self._export_history_to_csv(history, filepath, user_id)
            else:
                raise ValueError(f"Unsupported export format: {format_type}")
                
        except Exception as e:
            logger.error(f"Error exporting user history: {e}")
            raise
    
    def create_export_package(
        self,
        analysis_data: Dict[str, Any],
        user_id: str,
        include_formats: List[str] = ["pdf", "excel", "json"]
    ) -> str:
        """Create a zip package with multiple export formats"""
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            zip_filename = f"analysis_package_{timestamp}.zip"
            zip_filepath = self.export_dir / zip_filename
            
            with zipfile.ZipFile(zip_filepath, 'w', zipfile.ZIP_DEFLATED) as zipf:
                # Add different format exports
                for format_type in include_formats:
                    try:
                        if format_type == "pdf":
                            pdf_path = self.export_analysis_to_pdf(analysis_data, user_id)
                            zipf.write(pdf_path, f"analysis_report.{format_type}")
                        elif format_type == "excel":
                            excel_path = self.export_analysis_to_excel(analysis_data, user_id)
                            zipf.write(excel_path, f"analysis_report.{format_type}")
                        elif format_type == "json":
                            json_path = self._export_analysis_to_json(analysis_data, user_id)
                            zipf.write(json_path, f"analysis_report.{format_type}")
                    except Exception as e:
                        logger.warning(f"Could not add {format_type} to export package: {e}")
                
                # Add README file
                readme_content = f"""
Financial Document Analysis Export Package
Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
User ID: {user_id}

This package contains the analysis results in multiple formats:
- PDF: Formatted report suitable for presentation
- Excel: Structured data for further analysis
- JSON: Raw data for programmatic use

For questions or support, please contact the system administrator.
"""
                zipf.writestr("README.txt", readme_content)
            
            # Log export activity
            self.logging_service.log_activity(
                level=LogLevel.INFO,
                category=LogCategory.EXPORT,
                action=LogAction.EXPORT,
                message=f"Analysis package created: {zip_filename}",
                user_id=user_id,
                details={
                    "export_type": "package",
                    "filename": zip_filename,
                    "formats": include_formats,
                    "file_size": zip_filepath.stat().st_size
                }
            )
            
            return str(zip_filepath)
            
        except Exception as e:
            logger.error(f"Error creating export package: {e}")
            raise
    
    def _style_excel_sheet(self, sheet, title: str):
        """Apply styling to Excel sheet"""
        try:
            # Title styling
            title_cell = sheet['A1']
            title_cell.font = Font(bold=True, size=14, color="FFFFFF")
            title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            title_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Header styling
            for row in sheet.iter_rows(min_row=2, max_row=2):
                for cell in row:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
            
            # Auto-adjust column widths
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            logger.warning(f"Could not apply Excel styling: {e}")
    
    def _export_dashboard_to_excel(self, data: Dict[str, Any], filepath: Path, user_id: str) -> str:
        """Export dashboard data to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Dashboard Data"
        
        # Add dashboard data
        ws.append(['Dashboard Data Export'])
        ws.append(['Generated:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        ws.append([''])
        
        # Add various dashboard metrics
        for key, value in data.items():
            ws.append([key, str(value)])
        
        self._style_excel_sheet(ws, "Dashboard Data")
        wb.save(str(filepath))
        
        return str(filepath)
    
    def _export_dashboard_to_csv(self, data: Dict[str, Any], filepath: Path, user_id: str) -> str:
        """Export dashboard data to CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(['Metric', 'Value'])
            for key, value in data.items():
                writer.writerow([key, str(value)])
        
        return str(filepath)
    
    def _export_dashboard_to_json(self, data: Dict[str, Any], filepath: Path, user_id: str) -> str:
        """Export dashboard data to JSON"""
        export_data = {
            'export_info': {
                'generated_at': datetime.now().isoformat(),
                'user_id': user_id,
                'export_type': 'dashboard_data'
            },
            'data': data
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, default=str)
        
        return str(filepath)
    
    def _export_history_to_excel(self, history: List[Dict], filepath: Path, user_id: str) -> str:
        """Export user history to Excel"""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Analysis History"
        
        # Headers
        headers = ['Date', 'Document Name', 'Status', 'Analysis Type', 'Risk Level', 'Recommendation']
        ws.append(headers)
        
        # Data rows
        for record in history:
            row = [
                record.get('created_at', ''),
                record.get('filename', ''),
                record.get('status', ''),
                record.get('analysis_type', ''),
                record.get('risk_level', ''),
                record.get('recommendation', '')
            ]
            ws.append(row)
        
        self._style_excel_sheet(ws, "Analysis History")
        wb.save(str(filepath))
        
        return str(filepath)
    
    def _export_history_to_csv(self, history: List[Dict], filepath: Path, user_id: str) -> str:
        """Export user history to CSV"""
        with open(filepath, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Date', 'Document Name', 'Status', 'Analysis Type', 'Risk Level', 'Recommendation']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for record in history:
                writer.writerow({
                    'Date': record.get('created_at', ''),
                    'Document Name': record.get('filename', ''),
                    'Status': record.get('status', ''),
                    'Analysis Type': record.get('analysis_type', ''),
                    'Risk Level': record.get('risk_level', ''),
                    'Recommendation': record.get('recommendation', '')
                })
        
        return str(filepath)
    
    def _export_analysis_to_json(self, analysis_data: Dict[str, Any], user_id: str) -> str:
        """Export analysis to JSON format"""
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"analysis_report_{timestamp}.json"
        filepath = self.export_dir / filename
        
        export_data = {
            'export_info': {
                'generated_at': datetime.now().isoformat(),
                'user_id': user_id,
                'export_type': 'analysis_report'
            },
            'analysis_data': analysis_data
        }
        
        with open(filepath, 'w', encoding='utf-8') as jsonfile:
            json.dump(export_data, jsonfile, indent=2, default=str)
        
        return str(filepath)
    
    def cleanup_old_exports(self, days: int = 7):
        """Clean up export files older than specified days"""
        try:
            cutoff_date = datetime.now().timestamp() - (days * 24 * 60 * 60)
            cleaned_count = 0
            
            for file_path in self.export_dir.iterdir():
                if file_path.is_file() and file_path.stat().st_mtime < cutoff_date:
                    file_path.unlink()
                    cleaned_count += 1
            
            logger.info(f"Cleaned up {cleaned_count} old export files")
            return cleaned_count
            
        except Exception as e:
            logger.error(f"Error cleaning up old exports: {e}")
            return 0

# Global export service instance
export_service = None

def get_export_service() -> ExportService:
    """Get or create export service instance"""
    global export_service
    if export_service is None:
        export_service = ExportService()
    return export_service
